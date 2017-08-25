#!/usr/bin/env python2.7

from datetime import datetime, date, timedelta
import os
import sys
import re
import logging as mylog
import csv
import argparse
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import ColorScale, FormatObject
from openpyxl.styles import Color
from openpyxl.formatting.rule import Rule
import pandas



class Spreadsheeter():
    """Main class """

    def __init__(self, workbook, csv_files, append, headers, heatmap):
        self.header = headers
        self.csv_files = csv_files
        self.append = append
        self.workbook = self.open_workbook(workbook)
        self.workbook_name = workbook
        self.heatmap = heatmap

    def read_csv(self, csv_file):
        """ Read data from a csv file """
        mylog.debug('Reading csv file %s for data' % csv_file)
        csv_data = pandas.read_csv(csv_file)
        mylog.debug('Read of csv file complete.')
        return csv_data

    def append_csv_data(self, sheet, csv_data, header):
        """ This function will append data to a csv instead of overwrite it """
        mylog.debug('Appending csv data to sheet %s of %s' % (sheet, self.workbook))
        for row in dataframe_to_rows(csv_data, index=False, header=header):
            sheet.append(row)
            mylog.debug('Appended row %s' % row)
        mylog.debug('Completed appending csv data to sheet %s of %s' % (sheet, self.workbook))
        return True

    def integer_test(self, value):
        try:
            value = int(value)
            return True
        except:
            return False

    def float_test(self, value):
        try:
            value = float(value)
            return True
        except:
            return False

    def overwrite_sheet_with_csv_data(self, sheet, csv_data, header):
        """ openpyxl cannot overwrite sheet data. Only solution is to delete
        the sheet and replace it with a new one of the same name """
        sheet_name = sheet.title
        mylog.debug('Replacing csv data from sheet to %s in %s' %
                    (sheet_name, self.workbook_name))
        mylog.debug('Sheet name is %s' % sheet_name)
        self.workbook.remove_sheet(sheet)
        mylog.debug('Deleted sheet %s' % sheet)
        self.workbook.create_sheet(sheet_name)
        mylog.debug('Created sheet %s' % sheet_name)
        sheet = self.workbook.get_sheet_by_name(sheet_name)
        sheet_header = self.find_header(sheet_name)
        sheet.append(sheet_header)

        ## Some sheets have extra rows that need to be added ##
        ave_row = self.arsenal_v_everton(sheet_name)
        if ave_row:
            sheet.append(ave_row)
            mylog.debug('Adding special row %s to %s' % (ave_row, sheet_name))
        avm_row = self.arsenal_v_man(sheet_name)
        if avm_row:
            sheet.append(avm_row)
            mylog.debug('Adding special row %s to %s' % (avm_row, sheet_name))
        fcb_row = self.fcb_camp(sheet_name)
        if fcb_row:
            sheet.append(fcb_row)
            mylog.debug('Adding special row %s to %s' % (fcb_row, sheet_name))
        extra_row = self.extra_row(sheet_name)
        if extra_row:
            sheet.append(extra_row)
        blank_rows = self.blank_data(sheet_name)
        if blank_rows:
            for key, row in blank_rows.iteritems():
                mylog.debug('Adding blank row %s to %s' % (row, sheet_name))
                sheet.append(row)
        mylog.debug('Added header %s to sheet %s' % (sheet_header, sheet_name))
        for row in dataframe_to_rows(csv_data, index=False, header=header):
            ## For some reason some rows come out with the integers as strings ##
            ## This will erase such nonsense ##
            new_row = []
            for value in row:
                if type(value) is str:
                    int_test = self.integer_test(value)
                    if int_test == True:
                        value = int(value)
                        mylog.debug('Converted %s to integer' % value)
                    if int_test == False:
                        float_test = self.float_test(value)
                        if float_test == True:
                            value = float(value)
                            mylog.debug('Converted %s to float' % value)
                new_row.append(value)
            sheet.append(new_row)
            mylog.debug('Wrote row %s' % new_row)
        mylog.debug('Completed writing data from sheet to %s in %s' % (sheet_name, self.workbook_name))
        return True

    def date_range(self, start):
        """ Will generate a range of dates starting from the input start date
        to yesterday """
        current = start
        end = date.today()
        mylog.debug('Generating date list from %s to %s' % (start, end))
        while current < end:
            yield current
            current += timedelta(days=1)

    def blank_data(self, sheet):
        """ This will generate blank rows for a range of dates to be inserted
        into whichever sheet requires it.  This is because sometimes the DB has
        no entries for a particular date, and so does not supply any data, which
        throws off the pivot tables in the workbook """
        blank_rows = {}
        if sheet == 'data-cat_by_channel':
            mylog.debug('Found sheet requiring blank data, %s' % sheet)
            row_id = 1
            for dte in self.date_range(date(2017, 07, 21)):
                blank_row_list = [' ' + str(dte), 'Unassociated', 'Unassociated', 'Unassociated', 0]
                blank_rows[str(row_id)] = blank_row_list
                row_id += 1
        elif sheet == 'data-reg_by_channel':
            mylog.debug('Found sheet requiring blank data, %s' % sheet)
            row_id = 1
            for dte in self.date_range(date(2017, 07, 21)):
                blank_row_list = [' ' + str(dte), 'Unassociated', 'Unassociated', 0]
                blank_rows[str(row_id)] = blank_row_list
                row_id += 1
        elif sheet == 'data-reg_by_category':
            mylog.debug('Found sheet requiring blank data, %s' % sheet)
            row_id = 1
            for dte in self.date_range(date(2017, 07, 26)):
                blank_row_list = [' ' + str(dte), 'Unassociated', 0]
                blank_rows[str(row_id)] = blank_row_list
                row_id += 1
        elif sheet == 'data-site_reg_by_day':
            mylog.debug('Found sheet requiring blank data, %s' % sheet)
            row_id = 1
            for dte in self.date_range(date(2017, 8, 8)):
                blank_row_list = [' ' + str(dte), 0]
                blank_rows[str(row_id)] = blank_row_list
                row_id += 1
        else:
            blank_rows = None
        return blank_rows

    def extra_row(self, sheet):
        """ Some sheets have data that is not in the database and needs to be
        added individually """
        if sheet == 'data-site_reg_by_day':
            extra_row = [' 2017-05-05', 15]
        elif sheet == 'data-reg_by_channel':
            extra_row = [' 2017-06-06', '', '', 0]
        elif sheet == 'data-reg_by_category':
            extra_row = [' 2017-06-06', '', 0]
        else:
            extra_row = None
        return extra_row

    def arsenal_v_everton(self, sheet):
        if sheet == 'data-site_reg_by_day':
            a_v_e_row = ['Win a Trip to Arsenal v Everton', 10593]
        elif sheet == 'data-reg_by_channel' or sheet == 'data-imp_by_channel':
            a_v_e_row = ['Win a Trip to Arsenal v Everton', '', '', 10593]
        elif sheet == 'data-reg_by_category':
            a_v_e_row = ['Win a Trip to Arsenal v Everton', '', 10593]
        else:
            a_v_e_row = None
        return a_v_e_row

    def arsenal_v_man(self, sheet):
        if sheet == 'data-site_reg_by_day':
            a_v_m_row = ['Win a Trip to Arsenal v Man Utd', 218]
        elif sheet == 'data-reg_by_channel' or sheet == 'data-imp_by_channel':
            a_v_m_row = ['Win a Trip to Arsenal v Man Utd', '', '', 218]
        elif sheet == 'data-reg_by_category':
            a_v_m_row = ['Win a Trip to Arsenal v Man Utd', '', 218]
        else:
            a_v_m_row = None
        return a_v_m_row

    def fcb_camp(self, sheet):
        if sheet == 'data-site_reg_by_day':
            fcb_row = ['FCB Heilongjiang Camp - Nominate Yourself', 23469]
        elif sheet == 'data-reg_by_channel' or sheet == 'data-imp_by_channel':
            fcb_row = ['FCB Heilongjiang Camp - Nominate Yourself', '', '', 23469]
        elif sheet == 'data-reg_by_category':
            fcb_row = ['FCB Heilongjiang Camp - Nominate Yourself', '', 23469]
        else:
            fcb_row = None
        return fcb_row

    def find_header(self, sheet_name):
        """The sheets copied to from their respective csv files need a header
        in the top row so the pivot tables in other areas of the workbook can
        reference them properly.  This function adds the proper header to the
        respective sheet """
        header = []
        if sheet_name == 'data-reg_by_category':
            header = ['date', 'category', 'registrations']
        if sheet_name == 'data-reg_by_channel':
            header = ['Date', 'Campaign', 'Channel', 'Site Registrations']
        if sheet_name == 'data-reg_by_hour':
            header = ['Date', 'Hour', 'Percent', 'Count', 'Total']
        if sheet_name == 'data-imp_by_channel':
            header = ['Date', 'Campaign', 'Channel', 'Impressions']
        if sheet_name == 'data-cat_by_channel':
            header = ['Date', 'Category', 'Campaign', 'Channel', 'Count']
        if sheet_name == 'data-site_reg_by_day':
            header = ['Day', 'Registrations']
        if sheet_name == 'data-ctc_voting_activity':
            header = ['Date', 'Registrations']
        if sheet_name == 'data-site_reg_by_club':
            header = ['Club', 'Registrations']
        if sheet_name == 'data-direct_ms_reg_by_club':
            header = ['Club', 'Registrations']
        mylog.debug('Created header %s for sheet %s' % (header, sheet_name))
        return header

    def clean_csv(self, csv_file):
        """ This removes the non-csv lines from the top of the csv files
        generated by the database so they can be loaded properly by pandas """
        mylog.debug('Removing single or zero-data row from csv file %s' % csv_file)
        with open(csv_file, 'rb') as inp, open ('tmp_csv', 'wb') as out:
            writer = csv.writer(out)
            for row in csv.reader(inp):
                if len(row) < 2:
                    mylog.debug('Not re-writing row with data %s' % row)
                if len(row) >= 2:
                    writer.writerow(row)
        mylog.debug('Completed re-writing data to %s' % csv_file)
        os.rename('tmp_csv', csv_file)
        return True

    def open_workbook(self, workbook):
        """ Creates the workbook object """
        mylog.debug('Opening workbook %s' % workbook)
        workbook = openpyxl.load_workbook(filename = workbook)
        return workbook

    def add_color_conditioning(self, sheet, address_range):
        """ This will create  and apply a rule for color conditioning the heatmap sheet
        """
        mylog.debug('Begin add color conditioning to heatmap sheet for range %s' % address_range)
        mylog.debug('Setting minimum and maximum ranges, and mid to percentile with value of 50')
        first = FormatObject(type='min')
        mid = FormatObject(type='percentile', val=50)
        last = FormatObject(type='max')
        colors = [Color('7FA9D3'), Color('FCBF04'), Color('BF0200')]
        mylog.debug('Colors in range are %s' % colors)
        color_scale = ColorScale(cfvo=[first, mid, last], color=colors)
        rule = Rule(type='colorScale', colorScale=color_scale)
        mylog.debug('Applying conditional formatting to sheet at range %s' %
                address_range)
        sheet.conditional_formatting.add(address_range, rule)
        return True

    def get_heatmap_range(self, sheet):
        """  Find the range of the heatmap table and retrun the Grand Total
             column, as well as the column just before it.  To be used with
             the conditional formatting function """
        mylog.debug('Getting table size ranges for heatmap sheet table')
        for cell in sheet.get_cell_collection():
            if cell.value:
                if cell.value == 'Grand Total':
                    if cell.col_idx > 4:
                        total_column_idx = cell.col_idx
        table_end_column_idx = total_column_idx - 1
        total_column_ltr = openpyxl.utils.get_column_letter(total_column_idx)
        table_end_column_ltr = openpyxl.utils.get_column_letter(table_end_column_idx)
        table_range = 'C4:' + table_end_column_ltr + '27'
        total_range = total_column_ltr + '4:' + total_column_ltr + '27'
        mylog.debug('Sheet ranges are - table: %s, totals column: %s' %
                (table_range, total_range))
        return table_range, total_range

    def Execute(self):
        """ Main execution """
        mylog.info('Beginnig spreadhseet operations against %s' %
                   self.workbook_name)
        for csv_file in self.csv_files:
            mylog.debug('csv_file path is %s' % csv_file)
            mylog.info('Cleaning csv file %s' % csv_file)
            self.clean_csv(csv_file)
            csv_file_basename = os.path.basename(csv_file)
            sheet_name = os.path.splitext(csv_file_basename)[0]
            mylog.debug('Identified sheet to append %s data to is %s' % (csv_file_basename, sheet_name))
            csv_data = self.read_csv(csv_file)
            sheet = self.workbook.get_sheet_by_name(sheet_name)
            if self.append:
                mylog.info('Appending csv files %s to workbook %s' % (csv_file,
                    self.workbook_name))
                self.append_csv_data(sheet, csv_data, self.header)
            else:
                mylog.info('Writing csv files %s to workbook %s' % (csv_file,
                                                                    self.workbook_name))
                self.overwrite_sheet_with_csv_data(sheet, csv_data, self.header)
        if self.heatmap:
            mylog.info('Getting ranges of heatmap sheet table')
            sheet = self.workbook.get_sheet_by_name('Hourly Registration Heatmap')
            table_range, total_range = self.get_heatmap_range(sheet)
            mylog.info('Applying color conditioning to heatmap sheet tables')
            self.add_color_conditioning(sheet, table_range)
            self.add_color_conditioning(sheet, total_range)
        mylog.info('Saving changes to workbook %s' % self.workbook_name)
        self.workbook.save(self.workbook_name)


if __name__ == '__main__':
    #Parse command line arguments
    parser = argparse.ArgumentParser(add_help=True, description='Modify the BNN\
                                     workbook by importing csv data, adjusting\
                                     daily tables, and adding top 10 websites')
    parser.add_argument("-b", "--workbook", action="store", dest="workbook",\
                        default=None, help="workbook file to be modified")
    parser.add_argument("-a", "--append_csv", action="store", nargs='+',
                        dest="csv_append", default=None, help="list of csv files to APPEND to workbook")
    parser.add_argument("-p", "--csv_path", action="store", dest="csv_path",\
                        default='', help="path to csv files; include trailing /")
    parser.add_argument("-i", "--import_csv", action="store", nargs='+', dest="csv_import",
                        default=None, help="list of csv files to IMPORT to workbook")
    parser.add_argument("-m", "--heatmap", action="store_true", dest="heatmap", default=False, help = "just apply heatmap rules to workbook")
    parser.add_argument("--headers", action="store_true", dest="headers",\
                        default=True, help="if the csv files have headers that\
                        are not part of the dataset, indicate by using this option")
    parser.add_argument("--top_ten", action="store_true", dest="top_ten",
                        default=False, help="add the top ten sites to the workbook - not yet active")
    parser.add_argument("--debug", action="store_true", dest="debug",
                        default=False, help="set logging level to output debug messages")

    args = parser.parse_args()
    debug = args.debug

    if debug:
        mylog.basicConfig(stream=sys.stdout, level=mylog.DEBUG, format='%(asctime)s %(levelname)s - %(message)s')
    else: mylog.basicConfig(stream=sys.stdout, level=mylog.INFO, format='%(asctime)s %(levelname)s - %(message)s')
    if args.csv_append:
        append = True
    else:
        append = False
    if not args.workbook:
        print "You must supply a workbook upon which to operate."
        sys.exit(1)

    if args.top_ten:
        print "Adding top ten pages functionality unavailable at this time."

    # These are all of the known csv spreadsheets as of 6/28/17 #
    if not args.csv_append and not args.csv_import:
        csv_files = []
        csv_files.extend(('data-reg_by_hour.csv','data-site_reg_by_day.csv','data-reg_by_channel.csv',
                              'data-imp_by_channel.csv','data-reg_by_category.csv','data-cat_by_channel.csv',
                              'data-ctc_voting_activity.csv','data-site_reg_by_club.csv',
                              'data-direct_ms_reg_by_club.csv'))
    if args.csv_append:
        csv_files = args.csv_append
    if args.csv_import:
        csv_files = args.csv_import
    for idx, item in enumerate(csv_files):
        item = args.csv_path + item
        csv_files[idx] = item
    if args.heatmap:
        csv_files = []


    ModifySpreadsheet = Spreadsheeter(args.workbook, csv_files, append,
            args.headers, args.heatmap)
    try:
        ModifySpreadsheet.Execute()
    except SystemExit:
        print "System exit"
        raise
    except KeyboardInterrupt:
        print "Operation cancelled by user interrupt"
        sys.exit(1)
    except:
        print "Unknown exception:"
        raise
        sys.exit(1)


